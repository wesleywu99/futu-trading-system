"""Automated strategy optimization: find best single/combined strategy per stock.

Runs four phases:

1. Single-strategy baseline (5 strategies x 7 stocks x 3 periods)
2. Voting combinations   (C(5,2)=10 pairs x 7 stocks x 3 periods)
3. Entry/exit split      (5x4=20 pairs x 7 stocks x 3 periods)
4. Sequential combos     (10 pairs x 7 stocks x 1 period)

Each result is ranked by Sharpe/Sortino/win-rate/drawdown.  Final output:

* ``optimizer_results.xlsx`` — every backtest, ranked
* Per-stock Top 5 strategy recommendation printed to console

Filter for actionable recommendations:

* Sharpe > 0.5
* trades >= 10
* max drawdown < 30%
* stable across 3y and 5y windows (top 50%)

Usage:
    python tools/strategy_optimizer.py
"""

import os
import sys
import itertools
import time
from dataclasses import asdict

# Resolve project root so `src.*` imports work from anywhere
_PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if _PROJECT_ROOT not in sys.path:
    sys.path.insert(0, _PROJECT_ROOT)

import yfinance as yf

from src.backtest.engine import BacktestEngine, BacktestResult
from src.strategy.combined import CombinedStrategy
from src.strategy.ma_crossover import MACrossover
from src.strategy.macd_trend import MACDTrend
from src.strategy.bbands_rsi import BBandsRSI
from src.strategy.kdj_macd import KDJMACD
from src.strategy.adx_macd import ADXMACD

# ------------------------------------------------------------------
# Universe
# ------------------------------------------------------------------
STOCKS = {
    "AAPL":  "Apple",
    "MSFT":  "Microsoft",
    "AMZN":  "Amazon",
    "NVDA":  "NVIDIA",
    "GOOGL": "Google/Alphabet",
    "TSLA":  "Tesla",
    "META":  "Meta Platforms",
}

END_DATE = "2026-06-07"

PERIODS = [
    ("3y", 3 * 252),
    ("5y", 5 * 252),
    ("10y", 10 * 252),
]

# ------------------------------------------------------------------
# Strategy catalog (single strategies with default configs)
# ------------------------------------------------------------------
STRATEGY_CONFIGS = {
    "MACrossover": {
        "cls": MACrossover,
        "config": {
            "short_period": 5, "long_period": 20,
            "rsi_filter": {"enabled": True, "period": 14, "mode": "level", "upper_limit": 60},
        },
    },
    "MACDTrend": {
        "cls": MACDTrend,
        "config": {
            "fast_period": 12, "slow_period": 26, "signal_period": 9,
            "trend_ema_period": 30,
        },
    },
    "BBandsRSI": {
        "cls": BBandsRSI,
        "config": {
            "bb_period": 20, "bb_std_mult": 2.0,
            "rsi_period": 14, "rsi_oversold": 30, "rsi_buy_max": 45,
            "rsi_overbought": 70,
        },
    },
    "KDJMACD": {
        "cls": KDJMACD,
        "config": {
            "kdj_period": 9, "kdj_smooth": 3,
            "macd_fast": 12, "macd_slow": 26, "macd_signal": 9,
            "j_sell_threshold": 0,
        },
    },
    "ADXMACD": {
        "cls": ADXMACD,
        "config": {
            "ema_short": 13, "ema_mid": 55, "ema_long": 89,
            "adx_period": 14, "adx_threshold": 30, "adx_rising_bars": 2,
            "macd_fast": 12, "macd_slow": 26, "macd_signal": 9,
        },
    },
}

POSITION_PCT = 0.5
INITIAL_CASH = 100000

# ------------------------------------------------------------------
# Data
# ------------------------------------------------------------------
def download_all(period_years_max=10):
    """Download up to 10 years of daily data per ticker."""
    print("\n[1] Downloading historical data...")
    data = {}
    for ticker, name in STOCKS.items():
        try:
            df = yf.download(ticker, period=f"{period_years_max}y", end=END_DATE,
                             auto_adjust=True, progress=False)
        except Exception as e:
            print(f"    ! {ticker} download failed: {e}")
            continue
        if df is None or df.empty:
            print(f"    ! {ticker}: no data")
            continue
        klines = []
        for idx, row in df.iterrows():
            close = row["Close"]
            high = row["High"]
            low = row["Low"]
            vol = row["Volume"]
            if hasattr(close, "item"):
                close = close.item()
                high = high.item()
                low = low.item()
                vol = vol.item()
            klines.append({
                "date": idx.strftime("%Y-%m-%d"),
                "close": float(close),
                "high": float(high),
                "low": float(low),
                "volume": int(vol) if vol == vol else 0,  # NaN guard
            })
        data[ticker] = klines
        print(f"    {ticker:<6} {len(klines):>5} bars ({klines[0]['date']} -> {klines[-1]['date']})")
    return data


def slice_period(klines, num_bars):
    """Return the last ``num_bars`` of ``klines`` (or all if shorter)."""
    if num_bars is None or num_bars >= len(klines):
        return klines
    return klines[-num_bars:]


# ------------------------------------------------------------------
# Strategy factories
# ------------------------------------------------------------------
def make_single(name):
    spec = STRATEGY_CONFIGS[name]
    return spec["cls"], dict(spec["config"])


def make_pair(name_a, name_b, mode):
    """Return a factory callable that produces a fresh CombinedStrategy.

    A new instance is created per backtest so internal state is isolated.
    """
    spec_a = STRATEGY_CONFIGS[name_a]
    spec_b = STRATEGY_CONFIGS[name_b]
    cls_a, cls_b = spec_a["cls"], spec_b["cls"]
    cfg_a, cfg_b = dict(spec_a["config"]), dict(spec_b["config"])

    label = f"{name_a}+{name_b} [{mode}]"
    name = f"Combo_{mode}_{name_a}_{name_b}"

    def factory(store):
        sub_a = cls_a(f"_{name_a}", dict(cfg_a), store)
        sub_b = cls_b(f"_{name_b}", dict(cfg_b), store)
        return CombinedStrategy(
            name=name, config={"enabled": True}, store=store,
            sub_strategies=[sub_a, sub_b], mode=mode, label=label,
        )

    return name, label, factory


# ------------------------------------------------------------------
# Single backtest helpers
# ------------------------------------------------------------------
def run_single_backtest(ticker, klines, strat_name):
    """Backtest one single strategy on one kline series."""
    cls, config = make_single(strat_name)
    engine = BacktestEngine(
        strategy_cls=cls, strategy_config=config, code=ticker,
        initial_cash=INITIAL_CASH, commission=0.001,
        position_pct=POSITION_PCT,
    )
    return engine.run(klines)


def run_combined_backtest(ticker, klines, factory):
    """Backtest one CombinedStrategy (freshly built) on one kline series."""
    # Build a throwaway store just to construct sub-strategies; engine.run()
    # will rebind strategy.store to its own BacktestStore.
    from src.data.store import MarketDataStore
    placeholder = MarketDataStore()
    strategy_obj = factory(placeholder)
    engine = BacktestEngine(
        code=ticker, initial_cash=INITIAL_CASH, commission=0.001,
        position_pct=POSITION_PCT, strategy_obj=strategy_obj,
    )
    return engine.run(klines)


def result_to_row(r, strat_label, ticker, period):
    """Flatten a BacktestResult into a dict row."""
    pf = r.profit_factor if r.profit_factor != float("inf") else 999.0
    return {
        "stock": ticker,
        "period": period,
        "strategy": strat_label,
        "total_return_pct": round(r.total_return_pct, 2),
        "annual_return_pct": round(r.annual_return_pct, 2),
        "max_drawdown_pct": round(r.max_drawdown_pct, 2),
        "sharpe": round(r.sharpe_ratio, 4),
        "sortino": round(r.sortino_ratio, 4),
        "num_trades": r.num_trades,
        "win_rate_pct": round(r.win_rate, 2),
        "profit_factor": round(pf, 2),
        "avg_trade_return_pct": round(r.avg_trade_return_pct, 2),
    }


# ------------------------------------------------------------------
# Main pipeline
# ------------------------------------------------------------------
def main():
    start_time = time.time()
    data = download_all(period_years_max=10)
    if not data:
        print("No data. Aborting.")
        return

    all_rows = []
    strategy_names = list(STRATEGY_CONFIGS.keys())
    pair_combos = list(itertools.combinations(strategy_names, 2))
    ordered_pairs = [(a, b) for a in strategy_names for b in strategy_names if a != b]

    # ---------- Phase 1: single strategies ----------
    print(f"\n[2] Phase 1: single strategies "
          f"({len(strategy_names)} x {len(STOCKS)} x {len(PERIODS)} "
          f"= {len(strategy_names) * len(STOCKS) * len(PERIODS)} runs)")
    count = 0
    for strat_name in strategy_names:
        for ticker, klines in data.items():
            for plabel, pbars in PERIODS:
                sliced = slice_period(klines, pbars)
                if len(sliced) < 60:
                    continue
                try:
                    r = run_single_backtest(ticker, sliced, strat_name)
                except Exception as e:
                    print(f"    ! {strat_name}/{ticker}/{plabel} failed: {e}")
                    continue
                row = result_to_row(r, strat_name, ticker, plabel)
                row["mode"] = "single"
                all_rows.append(row)
                count += 1
                if count % 25 == 0:
                    print(f"    ... {count} single runs done "
                          f"({time.time()-start_time:.0f}s elapsed)")
    print(f"    Phase 1 done: {count} runs")

    # ---------- Phase 2: voting combinations ----------
    print(f"\n[3] Phase 2: voting combinations "
          f"({len(pair_combos)} x {len(STOCKS)} x {len(PERIODS)} "
          f"= {len(pair_combos) * len(STOCKS) * len(PERIODS)} runs)")
    count = 0
    for a, b in pair_combos:
        name, label, factory = make_pair(a, b, CombinedStrategy.VOTING)
        for ticker, klines in data.items():
            for plabel, pbars in PERIODS:
                sliced = slice_period(klines, pbars)
                if len(sliced) < 60:
                    continue
                try:
                    r = run_combined_backtest(ticker, sliced, factory)
                except Exception as e:
                    print(f"    ! {label}/{ticker}/{plabel} failed: {e}")
                    continue
                row = result_to_row(r, label, ticker, plabel)
                row["mode"] = "voting"
                all_rows.append(row)
                count += 1
                if count % 50 == 0:
                    print(f"    ... {count} voting runs done "
                          f"({time.time()-start_time:.0f}s elapsed)")
    print(f"    Phase 2 done: {count} runs")

    # ---------- Phase 3: entry/exit split ----------
    print(f"\n[4] Phase 3: entry/exit combinations "
          f"({len(ordered_pairs)} x {len(STOCKS)} x {len(PERIODS)} "
          f"= {len(ordered_pairs) * len(STOCKS) * len(PERIODS)} runs)")
    count = 0
    for a, b in ordered_pairs:
        name, label, factory = make_pair(a, b, CombinedStrategy.ENTRY_EXIT)
        for ticker, klines in data.items():
            for plabel, pbars in PERIODS:
                sliced = slice_period(klines, pbars)
                if len(sliced) < 60:
                    continue
                try:
                    r = run_combined_backtest(ticker, sliced, factory)
                except Exception as e:
                    print(f"    ! {label}/{ticker}/{plabel} failed: {e}")
                    continue
                row = result_to_row(r, label, ticker, plabel)
                row["mode"] = "entry_exit"
                all_rows.append(row)
                count += 1
                if count % 50 == 0:
                    print(f"    ... {count} entry/exit runs done "
                          f"({time.time()-start_time:.0f}s elapsed)")
    print(f"    Phase 3 done: {count} runs")

    # ---------- Phase 4: sequential (3y only for speed) ----------
    seq_period = PERIODS[0]  # 3y
    print(f"\n[5] Phase 4: sequential combinations "
          f"({len(pair_combos)} x {len(STOCKS)} x 1 "
          f"= {len(pair_combos) * len(STOCKS)} runs)")
    count = 0
    for a, b in pair_combos:
        name, label, factory = make_pair(a, b, CombinedStrategy.SEQUENTIAL)
        for ticker, klines in data.items():
            sliced = slice_period(klines, seq_period[1])
            if len(sliced) < 60:
                continue
            try:
                r = run_combined_backtest(ticker, sliced, factory)
            except Exception as e:
                print(f"    ! {label}/{ticker} failed: {e}")
                continue
            row = result_to_row(r, label, ticker, seq_period[0])
            row["mode"] = "sequential"
            all_rows.append(row)
            count += 1
    print(f"    Phase 4 done: {count} runs")

    # ---------- Analyze & export ----------
    print(f"\n[6] Analyzing {len(all_rows)} results...")
    export_excel(all_rows)
    print_top5_per_stock(all_rows)
    print_stability_winners(all_rows)

    elapsed = time.time() - start_time
    print(f"\nDone. Total {len(all_rows)} backtests in {elapsed/60:.1f} min.")
    print(f"Excel: {os.path.join(_PROJECT_ROOT, 'optimizer_results.xlsx')}")


# ------------------------------------------------------------------
# Excel export
# ------------------------------------------------------------------
def export_excel(rows):
    """Write every row + per-stock ranking sheets."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    if not rows:
        print("No rows to export.")
        return

    wb = openpyxl.Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    headers = [
        "stock", "period", "strategy", "mode",
        "total_return_pct", "annual_return_pct", "max_drawdown_pct",
        "sharpe", "sortino", "num_trades", "win_rate_pct",
        "profit_factor", "avg_trade_return_pct",
    ]

    # Sheet 1: all results
    ws = wb.active
    ws.title = "All Results"
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    rows_sorted = sorted(rows, key=lambda r: r["sharpe"], reverse=True)
    for ri, row in enumerate(rows_sorted, 2):
        for ci, h in enumerate(headers, 1):
            ws.cell(row=ri, column=ci, value=row.get(h))
        # Color the sharpe cell
        s = row["sharpe"]
        sharpe_cell = ws.cell(row=ri, column=headers.index("sharpe") + 1)
        if s >= 1.0:
            sharpe_cell.fill = green_fill
        elif s >= 0.5:
            sharpe_cell.fill = yellow_fill
        else:
            sharpe_cell.fill = red_fill
    for c in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 16

    # Sheet 2+: per stock ranking on 5y period (most stable window)
    target_period = "5y"
    for ticker in STOCKS:
        stock_rows = [
            r for r in rows
            if r["stock"] == ticker and r["period"] == target_period
        ]
        if not stock_rows:
            continue
        stock_rows.sort(key=lambda r: r["sharpe"], reverse=True)
        ws2 = wb.create_sheet(title=ticker[:31])
        rank_headers = ["rank", "pass_filter"] + headers
        for c, h in enumerate(rank_headers, 1):
            cell = ws2.cell(row=1, column=c, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        for i, r in enumerate(stock_rows, 1):
            passes = (
                r["sharpe"] > 0.5
                and r["num_trades"] >= 10
                and r["max_drawdown_pct"] < 30
            )
            ws2.cell(row=i + 1, column=1, value=i)
            ws2.cell(row=i + 1, column=2, value="YES" if passes else "no")
            for ci, h in enumerate(headers, 3):
                ws2.cell(row=i + 1, column=ci, value=r.get(h))
            fill = green_fill if passes else None
            if fill:
                for c in range(1, len(rank_headers) + 1):
                    ws2.cell(row=i + 1, column=c).fill = fill
        for c in range(1, len(rank_headers) + 1):
            ws2.column_dimensions[get_column_letter(c)].width = 16

    out_path = os.path.join(_PROJECT_ROOT, "optimizer_results.xlsx")
    wb.save(out_path)


# ------------------------------------------------------------------
# Reports
# ------------------------------------------------------------------
def _passes_filter(r):
    return (
        r["sharpe"] > 0.5
        and r["num_trades"] >= 10
        and r["max_drawdown_pct"] < 30
    )


def print_top5_per_stock(rows):
    print("\n" + "=" * 100)
    print("TOP 5 STRATEGIES PER STOCK (5y window, ranked by Sharpe)")
    print("=" * 100)
    target = "5y"
    for ticker in STOCKS:
        stock_rows = [
            r for r in rows
            if r["stock"] == ticker and r["period"] == target
        ]
        if not stock_rows:
            continue
        stock_rows.sort(key=lambda r: r["sharpe"], reverse=True)
        top = stock_rows[:5]
        print(f"\n  {ticker}:")
        print(f"  {'#':>2} {'Strategy':<42} {'Mode':<11} {'Sharpe':>7} {'Sortino':>8} "
              f"{'AnnRet%':>8} {'MaxDD%':>7} {'Win%':>6} {'Trades':>6} {'Pass':>5}")
        for i, r in enumerate(top, 1):
            passes = "YES" if _passes_filter(r) else "-"
            print(f"  {i:>2} {r['strategy']:<42} {r['mode']:<11} "
                  f"{r['sharpe']:>7.3f} {r['sortino']:>8.3f} "
                  f"{r['annual_return_pct']:>7.2f}% {r['max_drawdown_pct']:>6.2f}% "
                  f"{r['win_rate_pct']:>5.1f}% {r['num_trades']:>6} {passes:>5}")


def print_stability_winners(rows):
    """Find strategies that rank in the top 50% on both 3y and 5y windows."""
    print("\n" + "=" * 100)
    print("STABILITY WINNERS: top 50% on BOTH 3y and 5y windows")
    print("=" * 100)

    for ticker in STOCKS:
        for period_label in ("3y", "5y"):
            period_rows = [
                r for r in rows
                if r["stock"] == ticker and r["period"] == period_label
            ]
            period_rows.sort(key=lambda r: r["sharpe"], reverse=True)
            cutoff = len(period_rows) // 2 or 1
            top_set = {r["strategy"] for r in period_rows[:cutoff]}
            if period_label == "3y":
                three_y = top_set
            else:
                five_y = top_set
                stable = three_y & five_y
                if stable:
                    print(f"\n  {ticker}: {len(stable)} stable strategies")
                    for s in sorted(stable):
                        match_5y = next(
                            (r for r in period_rows if r["strategy"] == s), None
                        )
                        if match_5y:
                            print(f"    - {s:<40} Sharpe(5y)={match_5y['sharpe']:.3f}  "
                                  f"Trades={match_5y['num_trades']}  "
                                  f"MaxDD={match_5y['max_drawdown_pct']:.1f}%")
                else:
                    print(f"\n  {ticker}: no strategies stable across 3y & 5y")


if __name__ == "__main__":
    main()
