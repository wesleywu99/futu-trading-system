"""Backtest runner: Long-term strategy validation across market cycles.

Runs each Magnificent Seven stock with its optimal strategy over the maximum
available data range (up to 25 years), testing alpha across multiple bull/bear
cycles: dot-com crash, financial crisis, mobile internet era, COVID, AI rally.

Usage:
    python -m src.backtest.run
"""

import sys
import os

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))

import yfinance as yf
from src.backtest.engine import BacktestEngine

from src.strategy.ma_crossover import MACrossover
from src.strategy.macd_trend import MACDTrend
from src.strategy.bbands_rsi import BBandsRSI
from src.strategy.kdj_macd import KDJMACD
from src.strategy.adx_macd import ADXMACD

# ── Stock list with per-stock maximum data ranges ──
STOCKS = {
    "AAPL":  {"name": "Apple",          "start": "2000-01-01"},
    "MSFT":  {"name": "Microsoft",      "start": "2000-01-01"},
    "AMZN":  {"name": "Amazon",         "start": "2000-01-01"},
    "NVDA":  {"name": "NVIDIA",         "start": "2000-01-01"},
    "GOOGL": {"name": "Google/Alphabet","start": "2004-08-19"},
    "TSLA":  {"name": "Tesla",          "start": "2010-06-29"},
    "META":  {"name": "Meta Platforms", "start": "2012-05-18"},
}

END_DATE = "2026-06-07"

# ── Optimal strategy per stock (from 10-year backtest) ──
STOCK_STRATEGIES = {
    "AAPL": [(MACrossover, {
        "short_period": 5, "long_period": 20,
        "rsi_filter": {"enabled": True, "period": 14, "mode": "level", "upper_limit": 60},
    })],
    "MSFT": [(BBandsRSI, {
        "bb_period": 20, "bb_std_mult": 2.0,
        "rsi_period": 14, "rsi_oversold": 30, "rsi_buy_max": 45,
        "rsi_overbought": 70,
    })],
    "NVDA": [(MACDTrend, {
        "fast_period": 12, "slow_period": 26, "signal_period": 9,
        "trend_ema_period": 30,
    })],
    "GOOGL": [(ADXMACD, {
        "ema_short": 13, "ema_mid": 55, "ema_long": 89,
        "adx_period": 14, "adx_threshold": 30, "adx_rising_bars": 2,
        "macd_fast": 12, "macd_slow": 26, "macd_signal": 9,
    })],
    "TSLA": [(MACDTrend, {
        "fast_period": 12, "slow_period": 26, "signal_period": 9,
        "trend_ema_period": 30,
    })],
    "AMZN": [(ADXMACD, {
        "ema_short": 13, "ema_mid": 55, "ema_long": 89,
        "adx_period": 14, "adx_threshold": 30, "adx_rising_bars": 2,
        "macd_fast": 12, "macd_slow": 26, "macd_signal": 9,
    })],
    "META": [(MACDTrend, {
        "fast_period": 12, "slow_period": 26, "signal_period": 9,
        "trend_ema_period": 30,
    })],
}

POSITION_SIZES = [0.2, 0.5, 0.8]

# ── Market cycle definitions for segmented analysis ──
MARKET_CYCLES = [
    ("Dot-com Crash",    "2000-01", "2002-10"),
    ("Recovery",         "2002-10", "2007-10"),
    ("Financial Crisis", "2007-10", "2009-03"),
    ("Bull Market",      "2009-03", "2020-02"),
    ("COVID Crash",      "2020-02", "2020-03"),
    ("Post-COVID",       "2020-03", "2022-01"),
    ("Bear Market",      "2022-01", "2022-10"),
    ("AI Rally",         "2022-10", "2026-06"),
]


def download_data(ticker, start, end):
    """Download historical data from yfinance and convert to kline format."""
    print(f"  Downloading {ticker} from {start}...")
    df = yf.download(ticker, start=start, end=end, auto_adjust=True)
    if df.empty:
        print(f"  WARNING: No data for {ticker}")
        return []
    klines = []
    for idx, row in df.iterrows():
        close_val = row["Close"]
        high_val = row["High"]
        low_val = row["Low"]
        vol_val = row["Volume"]
        if hasattr(close_val, "item"):
            close_val = close_val.item()
            high_val = high_val.item()
            low_val = low_val.item()
            vol_val = vol_val.item()
        klines.append({
            "date": idx.strftime("%Y-%m-%d"),
            "close": float(close_val),
            "high": float(high_val),
            "low": float(low_val),
            "volume": int(vol_val),
        })
    return klines


def compute_bh_stats(klines):
    """Compute Buy & Hold stats for a kline series."""
    if not klines:
        return {"ret": 0, "ann": 0, "max_dd": 0}
    sp, ep = klines[0]["close"], klines[-1]["close"]
    ret = (ep - sp) / sp * 100
    years = len(klines) / 252
    ann = ((ep / sp) ** (1 / years) - 1) * 100 if years > 0 else 0
    peak, max_dd = sp, 0
    for k in klines:
        if k["close"] > peak:
            peak = k["close"]
        dd = (peak - k["close"]) / peak * 100
        if dd > max_dd:
            max_dd = dd
    return {"ret": ret, "ann": ann, "max_dd": max_dd}


def analyze_cycles(klines, equity_curve, trades):
    """Analyze strategy performance across market cycles.

    Returns dict: cycle_name -> {strategy_return, bh_return, excess, max_dd}
    """
    if not klines or len(equity_curve) != len(klines):
        return {}

    results = {}
    for cycle_name, start_str, end_str in MARKET_CYCLES:
        start_prefix = start_str[:7]  # "YYYY-MM"
        end_prefix = end_str[:7]

        # Find bar indices for this cycle
        start_idx = None
        end_idx = None
        for i, k in enumerate(klines):
            d = k["date"][:7]
            if start_idx is None and d >= start_prefix:
                start_idx = i
            if d <= end_prefix:
                end_idx = i

        if start_idx is None or end_idx is None or end_idx <= start_idx:
            continue

        # Strategy return in this cycle
        eq_start = equity_curve[start_idx]
        eq_end = equity_curve[end_idx]
        strat_ret = (eq_end - eq_start) / eq_start * 100 if eq_start > 0 else 0

        # B&H return in this cycle
        bh_start = klines[start_idx]["close"]
        bh_end = klines[end_idx]["close"]
        bh_ret = (bh_end - bh_start) / bh_start * 100 if bh_start > 0 else 0

        # Max drawdown in this cycle
        peak = equity_curve[start_idx]
        max_dd = 0
        for i in range(start_idx, end_idx + 1):
            peak = max(peak, equity_curve[i])
            dd = (peak - equity_curve[i]) / peak * 100
            if dd > max_dd:
                max_dd = dd

        # Count trades in this cycle
        cycle_trades = [
            t for t in trades
            if t.entry_date[:7] >= start_prefix and t.entry_date[:7] <= end_prefix
        ]

        results[cycle_name] = {
            "strategy_return": strat_ret,
            "bh_return": bh_ret,
            "excess": strat_ret - bh_ret,
            "max_dd": max_dd,
            "trades": len(cycle_trades),
            "bars": end_idx - start_idx + 1,
        }

    return results


def run_backtests():
    """Run long-term backtests: 7 stocks x 1 optimal strategy x 3 position sizes."""
    print("=" * 100)
    print("  LONG-TERM BACKTEST: Magnificent Seven with Optimal Strategies")
    print(f"  Maximum available data per stock, ending {END_DATE}")
    print("=" * 100)

    # Download data
    print("\n  Downloading historical data...")
    data = {}
    for ticker, info in STOCKS.items():
        klines = download_data(ticker, info["start"], END_DATE)
        if klines:
            data[ticker] = klines
            print(f"    {ticker}: {len(klines)} bars ({klines[0]['date']} to {klines[-1]['date']})")

    if not data:
        print("ERROR: No data downloaded.")
        return

    # Run backtests
    all_results = {}
    cycle_data = {}

    for ticker, klines in data.items():
        strategies = STOCK_STRATEGIES.get(ticker, [])
        for pos_pct in POSITION_SIZES:
            for strategy_cls, config in strategies:
                strat_name = strategy_cls.__name__
                print(f"\n  >> {ticker} ({STOCKS[ticker]['name']}) | {strat_name} | pos={pos_pct*100:.0f}%")
                engine = BacktestEngine(
                    strategy_cls=strategy_cls,
                    strategy_config=config,
                    code=ticker,
                    initial_cash=100000,
                    commission=0.001,
                    position_pct=pos_pct,
                )
                result = engine.run(klines)
                key = (strat_name, ticker, pos_pct)
                all_results[key] = result
                print(f"     Return: {result.total_return_pct:>8.2f}% | Sharpe: {result.sharpe_ratio:.4f} | "
                      f"Trades: {result.num_trades} | WinRate: {result.win_rate:.1f}%")

                # Cycle analysis (only for one position size to avoid duplication)
                if pos_pct == 0.5:
                    cycles = analyze_cycles(klines, result.equity_curve, result.trades)
                    cycle_data[(strat_name, ticker)] = cycles

    # Export to Excel
    export_excel(all_results, cycle_data, data)

    # Print report
    print_report(all_results, cycle_data, data)


def export_excel(results, cycle_data, data):
    """Export backtest results to Excel with market cycle analysis."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # B&H benchmarks
    bh = {}
    for ticker, klines in data.items():
        bh[ticker] = compute_bh_stats(klines)

    def write_headers(ws, headers):
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

    def auto_width(ws, num_cols):
        for c in range(1, num_cols + 1):
            ws.column_dimensions[get_column_letter(c)].width = 16

    # ── Sheet 1: Summary (7 stocks x 3 positions = 21 rows) ──
    ws1 = wb.active
    ws1.title = "Summary"
    headers1 = [
        "Stock", "Company", "Strategy", "Data Start", "Years",
        "Position Size", "Total Return %", "Annual Return %", "Max Drawdown %",
        "Sharpe Ratio", "Sortino Ratio", "Trades", "Win Rate %",
        "Profit Factor", "Avg Trade Return %",
        "B&H Return %", "B&H MaxDD %", "DD Savings vs B&H %",
    ]
    write_headers(ws1, headers1)

    row = 2
    for ticker, klines in data.items():
        strategies = STOCK_STRATEGIES.get(ticker, [])
        for strategy_cls, _ in strategies:
            strat_name = strategy_cls.__name__
            years = len(klines) / 252
            data_start = klines[0]["date"]
            for pos_pct in POSITION_SIZES:
                key = (strat_name, ticker, pos_pct)
                if key not in results:
                    continue
                r = results[key]
                dd_savings = bh[ticker]["max_dd"] - r.max_drawdown_pct
                pf = r.profit_factor if r.profit_factor != float("inf") else 999
                values = [
                    ticker, STOCKS[ticker]["name"], r.strategy_name, data_start,
                    round(years, 1), f"{pos_pct*100:.0f}%",
                    round(r.total_return_pct, 2), round(r.annual_return_pct, 2),
                    round(r.max_drawdown_pct, 2),
                    round(r.sharpe_ratio, 4), round(r.sortino_ratio, 4),
                    r.num_trades, round(r.win_rate, 2),
                    round(pf, 2), round(r.avg_trade_return_pct, 2),
                    round(bh[ticker]["ret"], 2), round(bh[ticker]["max_dd"], 2),
                    round(dd_savings, 2),
                ]
                for c, v in enumerate(values, 1):
                    ws1.cell(row=row, column=c, value=v)
                row += 1
    auto_width(ws1, len(headers1))

    # ── Sheet 2: vs Buy & Hold ──
    ws2 = wb.create_sheet("vs B&H")
    headers2 = [
        "Stock", "Company", "Strategy", "Years",
        "Strategy Return %", "B&H Return %", "Excess Return %",
        "Strategy MaxDD %", "B&H MaxDD %", "DD Savings %",
        "Strategy Sharpe", "Strategy Trades", "Strategy Win Rate %",
        "Alpha Rating",
    ]
    write_headers(ws2, headers2)

    row = 2
    for ticker, klines in data.items():
        strategies = STOCK_STRATEGIES.get(ticker, [])
        for strategy_cls, _ in strategies:
            strat_name = strategy_cls.__name__
            # Use pos=50% as representative
            key = (strat_name, ticker, 0.5)
            if key not in results:
                continue
            r = results[key]
            years = len(klines) / 252
            excess = r.total_return_pct - bh[ticker]["ret"]
            dd_saved = bh[ticker]["max_dd"] - r.max_drawdown_pct
            pf = r.profit_factor if r.profit_factor != float("inf") else 999

            # Alpha rating
            score = 0
            if r.sharpe_ratio > 0.5: score += 2
            elif r.sharpe_ratio > 0.3: score += 1
            if dd_saved > 15: score += 2
            elif dd_saved > 10: score += 1
            if r.win_rate > 50: score += 1
            if r.num_trades >= 15: score += 1
            # Check scaling
            scales = []
            for pp in POSITION_SIZES:
                k = (strat_name, ticker, pp)
                if k in results:
                    scales.append(results[k].sharpe_ratio)
            if len(scales) >= 2 and scales[-1] > scales[0]:
                score += 2
            if score >= 6: rating = "STRONG a"
            elif score >= 4: rating = "MODERATE a"
            elif score >= 2: rating = "WEAK a"
            else: rating = "NO a"

            values = [
                ticker, STOCKS[ticker]["name"], strat_name, round(years, 1),
                round(r.total_return_pct, 2), round(bh[ticker]["ret"], 2),
                round(excess, 2),
                round(r.max_drawdown_pct, 2), round(bh[ticker]["max_dd"], 2),
                round(dd_saved, 2),
                round(r.sharpe_ratio, 4), r.num_trades, round(r.win_rate, 2),
                rating,
            ]
            for c, v in enumerate(values, 1):
                cell = ws2.cell(row=row, column=c, value=v)

            fill = green_fill if "STRONG" in rating else (yellow_fill if "MODERATE" in rating else red_fill)
            for c in range(1, len(values) + 1):
                ws2.cell(row=row, column=c).fill = fill
            row += 1
    auto_width(ws2, len(headers2))

    # ── Sheet 3: Trades Detail ──
    ws3 = wb.create_sheet("Trades")
    headers3 = [
        "Stock", "Strategy", "Position Size",
        "Entry Date", "Exit Date", "Entry Price", "Exit Price",
        "Shares", "P&L $", "P&L %", "Holding Days",
    ]
    write_headers(ws3, headers3)

    row = 2
    for ticker in data:
        strategies = STOCK_STRATEGIES.get(ticker, [])
        for strategy_cls, _ in strategies:
            strat_name = strategy_cls.__name__
            for pos_pct in POSITION_SIZES:
                key = (strat_name, ticker, pos_pct)
                if key not in results:
                    continue
                r = results[key]
                for t in r.trades:
                    from datetime import datetime as dt
                    try:
                        d_entry = dt.strptime(t.entry_date, "%Y-%m-%d")
                        d_exit = dt.strptime(t.exit_date, "%Y-%m-%d")
                        hold_days = (d_exit - d_entry).days
                    except Exception:
                        hold_days = 0
                    values = [
                        ticker, strat_name, f"{pos_pct*100:.0f}%",
                        t.entry_date, t.exit_date,
                        round(t.entry_price, 2), round(t.exit_price, 2),
                        t.shares, round(t.pnl, 2), round(t.pnl_pct, 2),
                        hold_days,
                    ]
                    for c, v in enumerate(values, 1):
                        ws3.cell(row=row, column=c, value=v)
                    row += 1
    auto_width(ws3, len(headers3))

    # ── Sheet 4: Market Cycles ──
    ws4 = wb.create_sheet("Market Cycles")
    headers4 = [
        "Stock", "Strategy", "Cycle", "Period",
        "Strategy Return %", "B&H Return %", "Excess %",
        "Strategy MaxDD %", "Trades in Cycle", "Bars",
    ]
    write_headers(ws4, headers4)

    row = 2
    for ticker in data:
        strategies = STOCK_STRATEGIES.get(ticker, [])
        for strategy_cls, _ in strategies:
            strat_name = strategy_cls.__name__
            cycles = cycle_data.get((strat_name, ticker), {})
            for cycle_name, stats in cycles.items():
                # Find period string
                period = ""
                for cn, s, e in MARKET_CYCLES:
                    if cn == cycle_name:
                        period = f"{s} to {e}"
                        break
                values = [
                    ticker, strat_name, cycle_name, period,
                    round(stats["strategy_return"], 2),
                    round(stats["bh_return"], 2),
                    round(stats["excess"], 2),
                    round(stats["max_dd"], 2),
                    stats["trades"],
                    stats["bars"],
                ]
                for c, v in enumerate(values, 1):
                    cell = ws4.cell(row=row, column=c, value=v)
                # Color excess column
                if stats["excess"] > 5:
                    ws4.cell(row=row, column=7).fill = green_fill
                elif stats["excess"] < -5:
                    ws4.cell(row=row, column=7).fill = red_fill
                else:
                    ws4.cell(row=row, column=7).fill = yellow_fill
                row += 1
    auto_width(ws4, len(headers4))

    # Save
    out_path = os.path.join(
        os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))),
        "backtest_longterm_results.xlsx"
    )
    wb.save(out_path)
    print(f"\n  Excel saved: {out_path}")


def print_report(results, cycle_data, data):
    """Print comprehensive long-term backtest analysis."""

    bh = {}
    for ticker, klines in data.items():
        bh[ticker] = compute_bh_stats(klines)

    print(f"\n{'='*130}")
    print("  LONG-TERM BACKTEST: Magnificent Seven with Optimal Strategies")
    print(f"  Maximum available data per stock, ending {END_DATE}")
    print(f"{'='*130}")

    # ── B&H benchmarks ──
    print("\n  Buy & Hold benchmarks:")
    print(f"  {'Stock':<6} {'Company':<18} {'Data From':<12} {'Years':>6} {'B&H Return':>12} {'Annual':>9} {'B&H MaxDD':>10}")
    print(f"  {'-'*80}")
    for ticker, klines in data.items():
        years = len(klines) / 252
        print(
            f"  {ticker:<6} {STOCKS[ticker]['name']:<18} {klines[0]['date']:<12} "
            f"{years:>5.1f}y {bh[ticker]['ret']:>10.1f}% "
            f"{bh[ticker]['ann']:>7.1f}% {bh[ticker]['max_dd']:>8.1f}%"
        )

    # ── Per-stock results ──
    print(f"\n  {'='*120}")
    print("  STRATEGY RESULTS (all position sizes)")
    print(f"  {'='*120}")

    for ticker, klines in data.items():
        strategies = STOCK_STRATEGIES.get(ticker, [])
        for strategy_cls, _ in strategies:
            strat_name = strategy_cls.__name__
            print(f"\n  {ticker} ({STOCKS[ticker]['name']}) — {strat_name} — {len(klines)} bars ({klines[0]['date']} to {klines[-1]['date']})")
            print(f"  {'Pos':>4} {'Return%':>10} {'Annual%':>10} {'MaxDD%':>9} {'Sharpe':>9} {'Trades':>7} {'WinRate%':>9} {'vs B&H DD':>11}")
            print(f"  {'-'*75}")
            for pos_pct in POSITION_SIZES:
                key = (strat_name, ticker, pos_pct)
                if key in results:
                    r = results[key]
                    dd_diff = bh[ticker]["max_dd"] - r.max_drawdown_pct
                    print(
                        f"  {pos_pct*100:>3.0f}% "
                        f"{r.total_return_pct:>9.2f}% "
                        f"{r.annual_return_pct:>9.2f}% "
                        f"{r.max_drawdown_pct:>8.2f}% "
                        f"{r.sharpe_ratio:>9.4f} "
                        f"{r.num_trades:>7} "
                        f"{r.win_rate:>8.2f}% "
                        f"{dd_diff:>+9.1f}%"
                    )

    # ── Alpha scaling ──
    print(f"\n  {'='*120}")
    print("  ALPHA SCALING: Does Sharpe improve with position size?")
    print(f"  {'='*120}")
    for ticker in data:
        strategies = STOCK_STRATEGIES.get(ticker, [])
        for strategy_cls, _ in strategies:
            strat_name = strategy_cls.__name__
            scales = []
            for pos_pct in POSITION_SIZES:
                key = (strat_name, ticker, pos_pct)
                if key in results:
                    scales.append((pos_pct, results[key].sharpe_ratio))
            if len(scales) >= 2:
                tag = "SCALING" if scales[-1][1] > scales[0][1] else "NO SCALE"
                line = " -> ".join(f"pos={p*100:.0f}%:Sharpe={s:.4f}" for p, s in scales)
                print(f"    {ticker:<6} {strat_name:<14} {line}  [{tag}]")

    # ── Market cycle analysis ──
    print(f"\n  {'='*120}")
    print("  MARKET CYCLE ANALYSIS: Strategy performance across different market environments")
    print(f"  {'='*120}")

    for ticker in data:
        strategies = STOCK_STRATEGIES.get(ticker, [])
        for strategy_cls, _ in strategies:
            strat_name = strategy_cls.__name__
            cycles = cycle_data.get((strat_name, ticker), {})
            if not cycles:
                continue

            print(f"\n  {ticker} ({STOCKS[ticker]['name']}) — {strat_name}:")
            print(f"  {'Cycle':<20} {'Strat Ret%':>11} {'B&H Ret%':>10} {'Excess%':>9} {'MaxDD%':>8} {'Trades':>7}")
            print(f"  {'-'*70}")
            for cycle_name, stats in cycles.items():
                excess_str = f"{stats['excess']:>+8.1f}%"
                print(
                    f"  {cycle_name:<20} "
                    f"{stats['strategy_return']:>10.1f}% "
                    f"{stats['bh_return']:>9.1f}% "
                    f"{excess_str} "
                    f"{stats['max_dd']:>7.1f}% "
                    f"{stats['trades']:>7}"
                )

    # ── Final summary ──
    print(f"\n  {'='*120}")
    print("  FINAL SUMMARY: Long-term optimal strategy per stock")
    print(f"  {'='*120}")
    print(f"  {'Stock':<6} {'Strategy':<14} {'Period':<24} {'Years':>6} {'Return%':>10} {'B&H%':>10} {'MaxDD%':>8} {'Sharpe':>8} {'Trades':>7} {'WinRate':>8}")
    print(f"  {'-'*110}")

    for ticker in data:
        strategies = STOCK_STRATEGIES.get(ticker, [])
        for strategy_cls, _ in strategies:
            strat_name = strategy_cls.__name__
            key = (strat_name, ticker, 0.8)  # Use 80% position for best case
            if key not in results:
                key = (strat_name, ticker, 0.5)
            if key not in results:
                continue
            r = results[key]
            klines = data[ticker]
            years = len(klines) / 252
            period = f"{klines[0]['date']} to {klines[-1]['date']}"
            print(
                f"  {ticker:<6} {strat_name:<14} {period:<24} "
                f"{years:>5.1f}y "
                f"{r.total_return_pct:>9.2f}% "
                f"{bh[ticker]['ret']:>9.1f}% "
                f"{r.max_drawdown_pct:>7.2f}% "
                f"{r.sharpe_ratio:>8.4f} "
                f"{r.num_trades:>7} "
                f"{r.win_rate:>7.1f}%"
            )

    print(f"\n{'='*130}")
    print("  NOTE: This long-term backtest covers multiple market cycles.")
    print("  Compare with the previous 10-year (2015-2025) results to assess alpha persistence.")
    print(f"{'='*130}")


if __name__ == "__main__":
    run_backtests()
